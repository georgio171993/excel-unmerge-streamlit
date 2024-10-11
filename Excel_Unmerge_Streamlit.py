import openpyxl
import streamlit as st
from datetime import datetime

# Function to process the Excel file
def process_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # 1. Perform the necessary merges based on the new conditions
    st.write("Merging rows under 1_3_1_1 and 1_3_3 if 1_2_3 is the same...")
    
    for row in range(2, ws.max_row + 1):  # Assuming first row is header
        # Column indexes: 1_2_3 -> 3rd column, 1_3_1_1 -> 4th column, 1_3_3 -> 6th column
        if ws.cell(row=row, column=12).value == ws.cell(row=row-1, column=12).value:  # 1_2_3 is the 3rd column
            # Merge 1_3_1_1 (Column 4) if values are the same
            if ws.cell(row=row, column=18).value == ws.cell(row=row-1, column=18).value:  # 1_3_1_1 is the 4th column
                ws.merge_cells(start_row=row-1, start_column=18, end_row=row, end_column=18)
            
            # Merge 1_3_3 (Column 6) if values are the same
            if ws.cell(row=row, column=20).value == ws.cell(row=row-1, column=20).value:  # 1_3_3 is the 6th column
                ws.merge_cells(start_row=row-1, start_column=20, end_row=row, end_column=20)

    # 2. Unmerge values in the specified columns
    unmerge_columns = [11, 12, 44, 45, 57]  # Columns for 1_2_2, 1_2_3, 1_4_2, 1_4_3, 1_5_4
    st.write("Unmerging cells in columns: 1_2_2, 1_2_3, 1_4_2, 1_4_3, 1_5_4...")
    
    for merge in list(ws.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        
        if min_col in unmerge_columns:
            ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    # 3. Replace "Metric Ton" with blank in specified columns (1_5_8, 1_5_9, 1_5_12)
    replace_columns = [62, 63, 68]  # Columns for 1_5_8, 1_5_9, 1_5_12
    st.write("Replacing 'Metric Ton' with blanks in columns: 1_5_8, 1_5_9, 1_5_12...")
    
    for row in ws.iter_rows():
        for col_idx in replace_columns:
            cell_value = row[col_idx-1].value  # Adjust column index (0-based)
            if isinstance(cell_value, str) and "Metric Ton" in cell_value:
                row[col_idx-1].value = cell_value.replace("Metric Ton", "").strip()

    # 4. Unmerge cells and handle the unmerged values (existing logic)
    st.write("Unmerging other cells and handling values...")
    
    for merge in list(ws.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        
        first_cell_value = ws.cell(row=min_row, column=min_col).value
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                ws.cell(row=row, column=col).value = None
    
    # 5. Condition to replace 'N/A' with blanks
    st.write("Replacing 'N/A' with blanks...")
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'N/A':
                cell.value = ''

    # Create output file name with the current date and time
    current_time = datetime.now().strftime('%d %b %Y %H:%M')
    output_file = f"Questionnaire Answers - {current_time}.xlsx"
    
    wb.save(output_file)
    return output_file

# Streamlit app
st.title("Excel Unmerge and Update Tool")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    st.write("Processing...")
    output_file = process_excel(uploaded_file)
    
    with open(output_file, "rb") as f:
        st.download_button("Download the processed file", f, file_name=output_file)
